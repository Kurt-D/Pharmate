#!/usr/bin/env python3
"""
Dev-only converter: Drug Reference.xlsx  ->  seeds/formulary_raw.json + interactions_raw.json

Produces a FAITHFUL structured dump of the pharmacist curation workbook.
All normalization (name cleanup, severity mapping, interaction-type derivation,
the unsigned-row guard) happens in the Node seed loader so it stays testable in
the app language. Re-run this whenever the pharmacist returns an updated workbook.

Usage (from repo root):
    python seeds/xlsx_to_json.py "Drug Reference.xlsx"
"""
import json
import sys
from pathlib import Path

import openpyxl


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    # Repair common mojibake from the source file's dash/quote encoding.
    s = s.replace("�", "-")
    return s if s != "" else None


def main():
    src = Path(sys.argv[1] if len(sys.argv) > 1 else "Drug Reference.xlsx")
    out_dir = Path(__file__).parent
    wb = openpyxl.load_workbook(src, data_only=True)

    # ─── Drug Reference sheet ─────────────────────────────────────────────────
    ref_ws = wb["Drug Reference"]
    rows = list(ref_ws.iter_rows(values_only=True))
    # The workbook has a duplicated header (rows 0 and 1); real data starts row 2.
    formulary = []
    for r in rows[2:]:
        if all(c is None for c in r):
            continue
        formulary.append({
            "generic_name": clean(r[0]),
            "brand_names": clean(r[1]),
            "standard_frequency": clean(r[2]),
            "min_interval_hours": clean(r[3]),
            "meal_instruction": clean(r[4]),
            "max_daily_doses": clean(r[5]),
            "is_prn_default": clean(r[6]),
            "default_interval_hours": clean(r[7]),
            "meal_anchor_code": clean(r[8]),
            "notes": clean(r[9]),
            "verified_by": clean(r[10]) if len(r) > 10 else None,
            "date_approved": clean(r[11]) if len(r) > 11 else None,
        })

    # ─── Drug Interaction Pairs sheet ─────────────────────────────────────────
    pair_ws = wb["Drug Interaction Pairs"]
    prows = list(pair_ws.iter_rows(values_only=True))
    interactions = []
    for r in prows[1:]:
        if all(c is None for c in r):
            continue
        interactions.append({
            "drug_1": clean(r[0]),
            "drug_2": clean(r[1]),
            "severity": clean(r[2]),
            "min_gap_hours": clean(r[3]),
            "note": clean(r[4]),
        })

    (out_dir / "formulary_raw.json").write_text(
        json.dumps(formulary, indent=2, ensure_ascii=False), encoding="utf-8")
    (out_dir / "interactions_raw.json").write_text(
        json.dumps(interactions, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"Wrote {len(formulary)} drugs -> formulary_raw.json")
    print(f"Wrote {len(interactions)} pairs -> interactions_raw.json")


if __name__ == "__main__":
    main()
